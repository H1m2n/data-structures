# uncensor("Wh*r* d*d my v*w*ls g*?", "eeioeo") ➞ "Where did my vowels go?"


def uncensored(sen, vowel):
    i = 0
    out = ''
    for x in sen:
        if x == '*':
            out += vowel[i]
            i += 1
        else:
            out += x
    return out


print(uncensored("Wh*r* d*d my v*w*ls g*?", "eeioeo"))


def num_func(num):
    if num % 3 == 0 and num % 5 == 0:
        return 'FizzBuzz'
    if num % 3 == 0:
        return "Fizz"
    if num % 5 == 0:
        return 'Buzz'

    return str(num)


print(num_func(15))
print(num_func(5))
print(num_func(3))

s = ['sa', 'is', 'si', 'palm', 'mpal', 'a', 'a', 'aplm', 'psam', 'mpas', 'pmas', 'pmsa', 'as']

s1 = sorted(s, key=lambda x: len(x))
print(s1)

i, j = (0, 1)
out = []
while j < len(s):
    l = ''.join(sorted(s1[i]))
    r = ''.join(sorted(s1[j]))
    if l == r:
        if len(out) > 0 and ''.join(sorted(out[-1][-1])) == r:
            out[-1].append(s1[j])
        else:
            out.append([s1[i], s1[j]])
    else:
        pass
    i += 1
    j += 1
print(out)


def arg_test(*args, **kwargs):
    print(args, '************')
    print(args[0], args[1])


arg_test(*(1, 2))
arg_test(3, 4)

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.create_sheet("Validation")


data = [('Sr_No', "", "CVE-ID", "Test Steps", "Expected Result(NVD)", "Actual Result(Spotlight)", "Status")]
data.append(('', '', 1, "Verify Description", 'description_nvd', 'description_spotlight', 'store'))
data.append(('', '', 1, "Verify CVSS Base Score", 'base_score_cvss_nvd', 'base_score_cvss_spotlight', 'store1'))
data.append(('', '', 2, "Verify CVSS Base Score", 'base_score_cvss_nvd', 'base_score_cvss_spotlight', 'store1'))

for row_i, row in enumerate(data):
    for col_i, col in enumerate(row):
        if col_i == 0 and row_i >= 1:
            col = row_i
        sheet.cell(row=row_i + 1, column=col_i + 1).value = col

wb.save('test-sheet.xlsx')


# import re
# re.search()



